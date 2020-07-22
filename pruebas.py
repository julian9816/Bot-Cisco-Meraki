from meraki_sdk import *
from meraki_sdk.meraki_sdk_client import MerakiSdkClient
from meraki_sdk.exceptions.api_exception import APIException
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font

x_cisco_meraki_api_key = "1833bcc16a027bf707548bdce8a978e7c517153e"
meraki = MerakiSdkClient(x_cisco_meraki_api_key)
organization_id={"id":"921480"}
networks=meraki.networks.get_organization_networks({"organization_id":"921480"})
print(networks)
# controladores
networkid = "L_635570497412679705"
ssid_controller = meraki.clients
collect = {}
collect["network_id"] = networkid
def as_text(value):
       if value is None:
           return ""
       return str(value)
def ajustarTamaño(filesheet):
        wb=openpyxl.load_workbook(filesheet)
        for n in wb.get_sheet_names():
             worksheet=wb.get_sheet_by_name(n)
             for column_cells in worksheet.columns:
                   length = max(len(as_text(cell.value)) for cell in column_cells)+2
                   worksheet.column_dimensions[column_cells[0].column_letter].width = length
                   for cell in column_cells:
                       cell.alignment = Alignment(horizontal='center')
        wb.save(filesheet)
try:
    ssid_controller.get_network_clients(collect)
    clients=ssid_controller.get_network_clients(collect)
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title="Clientes")
    hoja = wb['Clientes']
    descripciones=['Descripcion', 'Visto por primera vez', 'Group Policy 8021x',
        'ID', 'Direccion Ip', 'ip6', 'Direccion Ip V6 Local', 'Visto por ultima vez', 'Direccion MAC', 'Fabricante',
        'notes', 'Sistema Operativo', 'MAC Dispositivo de la Red', 'Dispositivo de la Red', 'Serial Dispositivo de la Red',
        'smInstalled', 'SSID', 'Estado', 'Puerto del Switch', 'Datos recibidos (Mb)', 'Datos enviados (Mb)', 'Usuarios', 'VLAN']
    hoja.append(descripciones)
    j=1
    for i in clients:
        valores=[]
        for desc, valor in i.items():
            if desc=="usage":
                valores.append(int(valor["recv"])/1000)
                valores.append(int(valor["sent"])/1000)
            else:
                valores.append(valor)
            descripciones.append(desc)
        hoja.append(valores)
        cell=hoja["T"+str(j+1)]
        cell.number_format='#,##0'
        cell=hoja["U"+str(j+1)]
        cell.number_format='#,##0'
        cell=hoja["H"+str(j+1)]
        fecha=str(cell.value)
        hoja["H"+str(j+1)]=(fecha.split("T")[0]+" || "+fecha.split("T")[1].split("Z")[0].split(":")[0]+":"+
                            fecha.split("T")[1].split("Z")[0].split(":")[1])
        cell=hoja["B"+str(j+1)]
        fecha=str(cell.value)
        hoja["B"+str(j+1)]=(fecha.split("T")[0]+" || "+fecha.split("T")[1].split("Z")[0].split(":")[0]+":"+
                            fecha.split("T")[1].split("Z")[0].split(":")[1])
        collect["client_id"]=i["id"]
        if i["description"]==None:
            wb.create_sheet(index=j, title=i["manufacturer"])
            hoja1=wb[i["manufacturer"]]
        else:
            wb.create_sheet(index=j, title=i["description"])
            hoja1=wb[i["description"]]
        descripciones=['Segundos Activo', 'Aplicacion', 'Destino', 'numFlows',
                        'Puerto', 'Protocolo', 'Datos recibidos (KB)', 'Datos Enviados (KB)', 'ts']
        hoja1.append(descripciones)
        trafico=ssid_controller.get_network_client_traffic_history(collect)
        for i in trafico:
            valores=[]
            for desc, valor in i.items():
                valores.append(valor)
            hoja1.append(valores)
            hoja1.delete_cols(9)
            for row in range(2,hoja1.max_row+1):
                for column in "GH":  #Here you can add or reduce the columns
                    cell_name = "{}{}".format(column, row)
                    cell=hoja1[cell_name]# the value of the specific cell
                    cell.number_format='#,##0'
        for column in "ABCDEFGHIJKLMNOPQRSTUV":
            cell_name = "{}{}".format(column, 1)
            hoja[cell_name ].font = Font(bold=True)
            hoja1[cell_name ].font = Font(bold=True)
        j+=1
    hoja.delete_cols(3)
    wb.save('clientes.xlsx')
    ajustarTamaño('clientes.xlsx')

except APIException as e:
    print(e)

