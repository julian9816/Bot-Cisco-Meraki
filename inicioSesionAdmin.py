import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import unittest
import os
import shutil
import glob
from datetime import datetime
import openpyxl 
import smtplib
import email
import openpyxl
import smtplib
import email
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from os.path import basename
from openpyxl.styles import Alignment

class inicioSesionAdmin(unittest.TestCase):

    def setUp(self):
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_experimental_option("prefs", {
          "download.default_directory": r"E:\ARUS S.A\Proyectos\Python\Webex\Chrome_Driver\Chrome_Driver\Archivos\Archivos",
          "download.prompt_for_download": False,
          "download.directory_upgrade": True,
          "safebrowsing.enabled": True
        })
        self.driver = webdriver.Chrome(chrome_options=chrome_options, executable_path=r'chromedriver.exe')
        self.correos=['1']
        self.filesheet=""
    def test_nombreUsuarios(self):
        correos=self.correos
        driver = self.driver
        driver.get('https://admin.webex.com/');
                            # Optional argument, if not specified will search path.
                         # Let the user actually see something!
        wait=WebDriverWait(driver, 50)
        wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/webex-root/webex-login/div/div/div[1]/div[2]/form/div[1]/div/input')))
        usuario=driver.find_element_by_xpath('/html/body/webex-root/webex-login/div/div/div[1]/div[2]/form/div[1]/div/input')
        correo='csjulian@unicauca.edu.co'
        usuario.send_keys(correo)
        usuario.send_keys(Keys.ENTER)
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="IDToken2"]')))
        clave = driver.find_element_by_xpath('//*[@id="IDToken2"]')
        clave.send_keys("G849jNN43#")
        clave.send_keys(Keys.ENTER)
        wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/webex-root/webex-main/nav/webex-sidebar/md-sidebar/div[1]/md-sidebar-header/img'))) # Let the user actually see something!
        driver.get('https://admin.webex.com/reports/webexMetrics/classic')
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="training_center"]/h3')))
        analisis=driver.find_element_by_xpath('//*[@id="training_usage"]')
        analisis.click()
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="webexIframeContainer"]')))
        #Cambio al iframe
        driver.switch_to_frame(driver.find_element_by_xpath('//*[@id="webexIframeContainer"]'))
        wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/form/table/tbody/tr[4]/td[2]/table/tbody/tr[2]/td[2]/select[1]')))
        fecha=datetime.now()
        #FECHA INICIO-REPORT
        mesInicio=Select(driver.find_element_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/form/table/tbody/tr[4]/td[2]/table/tbody/tr[2]/td[2]/select[2]'))
        mesInicio.select_by_value("4")
        diaInicio=Select(driver.find_element_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/form/table/tbody/tr[4]/td[2]/table/tbody/tr[2]/td[2]/select[1]'))
        diaInicio.select_by_value("1")
        #FECHA FIN-REPORT
        mesFin=Select(driver.find_element_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/form/table/tbody/tr[4]/td[2]/table/tbody/tr[3]/td[2]/select[2]'))
        mesFin.select_by_value(str(fecha.month))
        diaFin=Select(driver.find_element_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/form/table/tbody/tr[4]/td[2]/table/tbody/tr[3]/td[2]/select[1]'))
        diaFin.select_by_value(str(fecha.day))
        #Ordenar por Usuario
        orden=Select(driver.find_element_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/form/table/tbody/tr[4]/td[2]/table/tbody/tr[5]/td[2]/select'))
        orden.select_by_value("9")
        #boton display
        btnDisplay=driver.find_element_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td/form/table/tbody/tr[4]/td[2]/table/tbody/tr[6]/td[2]/input')
        btnDisplay.click()
        wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table/tbody/tr[5]/td/table')))
        #sacar los usuarios de cada profesor
        rows=len(driver.find_elements_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table/tbody/tr[5]/td/table/tbody/tr'))
        #descargar el reporte general
        informe=driver.find_elements_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table/tbody/tr[2]/td[2]/input[2]')
        informe[0].click()
        time.sleep(2)
        filepath1 = 'E:\ARUS S.A\Proyectos\Python\Webex\Chrome_Driver\Chrome_Driver\Archivos'
        filepath = 'E:\ARUS S.A\Proyectos\Python\Webex\Chrome_Driver\Chrome_Driver\Archivos\Archivos'
        filename = max([f for f in os.listdir(filepath)], key=lambda xa : os.path.getctime(os.path.join(filepath,xa)))
        if '.part' in filename:
             time.sleep(2)
             programa='excel ' + '"' +os.path.join(filepath, filename)+'" ' +'"'+ os.path.join(filepath1,'Reporte General '+str(fecha.date())+'.xlsx')+'"'
             os.system(programa)
        else:
            programa='excel ' + '"' +os.path.join(filepath, filename)+'" ' +'"'+ os.path.join(filepath1,'Reporte General '+str(fecha.date())+'.xlsx')+'"'
            os.system(programa)
        informeGeneral=os.path.join(filepath1,'Reporte General '+str(fecha.date())+'.xlsx')
        #organizar informe general
        workbook=openpyxl.load_workbook(informeGeneral)
        sheet_general=workbook.active
        sheet_general.delete_cols(1)
        sheet_general.delete_rows(1,3)
        sheet_general.delete_cols(9)
        sheet_general.delete_cols(11,5)
        sheet_general.delete_cols(11)
        sheet_general['B1'].value="Clase"
        sheet_general.title="Reporte General "+str(fecha.date())
        workbook.save(informeGeneral)
        os.remove(os.path.join(filepath,filename))
        ##########
        filesheet = 'E:\ARUS S.A\Proyectos\Python\Webex\Chrome_Driver\Chrome_Driver\Archivos\Reporte Especifico '+str(fecha.date())+'.xlsx'
        self.filesheet=filesheet
        self.fecha=str(fecha.date())
        wb=openpyxl.Workbook()
        wb.save(filesheet)
        #filesheet=os.path.join(filepath,"Reporte Total "+str(fecha.date())+".xlsx")
        i=0
        for n in range(0,rows-1):
            clase=driver.find_elements_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table/tbody/tr[5]/td/table/tbody/tr['+str(n+2)+']/td[1]/font/a')
            nombre_clase=clase[0].text
            invitados=driver.find_elements_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table/tbody/tr[5]/td/table/tbody/tr['+str(n+2)+']/td[6]/font')[0].text
            participantes=driver.find_elements_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table/tbody/tr[5]/td/table/tbody/tr['+str(n+2)+']/td[8]/font')[0].text
            ausentes=driver.find_elements_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table/tbody/tr[5]/td/table/tbody/tr['+str(n+2)+']/td[9]/font')[0].text
            correo1=driver.find_elements_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table/tbody/tr[5]/td/table/tbody/tr['+str(n+2)+']/td[2]/font/a')
            correo=correo1[0].text
            clase[0].click()
            wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table[1]/tbody/tr[2]/td[2]/input[2]')))
            descarga=driver.find_elements_by_xpath('/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table[1]/tbody/tr[2]/td[2]/input[2]')[0].click()
            time.sleep(2)
            filepath = 'E:\ARUS S.A\Proyectos\Python\Webex\Chrome_Driver\Chrome_Driver\Archivos\Archivos'
            
            filename = max([f for f in os.listdir(filepath)], key=lambda xa : os.path.getctime(os.path.join(filepath,xa)))
            newname=correo+' '+nombre_clase+'.csv'
 
            if '.part' in filename:
                time.sleep(2)
                os.rename(os.path.join(filepath, filename), os.path.join(filepath, newname))
            else:
                os.rename(os.path.join(filepath, filename),os.path.join(filepath,newname))
            programa='excel ' + '"' +os.path.join(filepath,newname) +'" ' +'"'+ os.path.join(filepath,correo+' '+nombre_clase+'.xlsx')+'"'
            os.system(programa)
            if correos[-1]==correo:
                wb=openpyxl.load_workbook(filesheet)
                sheet_Total=wb.get_sheet_by_name(correo)
                excel_document = openpyxl.load_workbook(os.path.join(filepath,correo+' '+nombre_clase+'.xlsx'))
                sheet=excel_document.active
                #Ele,mentos que toca borrar
                sheet.delete_cols(6,5)
                sheet.delete_cols(11)
                sheet.delete_cols(12,7)
                sheet.delete_rows(1)
                sheet.delete_cols(11,2)
                sheet.insert_rows(2,2)
                total_par=sheet[1]
                profesor=sheet[6]
                profesor[0].value="Profesor"
                total_par[0].value=nombre_clase
                total_par[1].value="Total Participantes"
                total_par[2].value=participantes
                total_par[2].value=int(participantes)-1
                total_aus=sheet[2]
                total_aus[1].value="Ausentes"
                total_aus[2].value=int(ausentes)
                total_est=sheet[3]
                total_est[1].value="Total Estudiantes"
                total_est[2].value=int(ausentes)+int(participantes)-1
                for n in range(1,int(invitados)):
                    sheet.cell(row = n+6, column = 1).value=n
                i=0
                inicio=sheet_Total.max_row+2
                for row in sheet.iter_rows(min_row=1,max_row=sheet.max_row,min_col=1, max_col=sheet.max_column):
                    j=0
                    fila=sheet_Total[inicio+i]
                    i=i+1
                    for cell in row:
                               fila[j].value=cell.value
                               if j==6:
                                   fila[j].number_format='mm-dd-yyyy'
                               j=j+1
                wb.save(filesheet)
            else:
                wb=openpyxl.load_workbook(filesheet)
                wb.create_sheet(correo)
                sheet_Total=wb.get_sheet_by_name(correo)
                excel_document = openpyxl.load_workbook(os.path.join(filepath,correo+' '+nombre_clase+'.xlsx'))
                sheet=excel_document.active
                sheet.delete_cols(6,5)
                sheet.delete_cols(11)
                sheet.delete_cols(12,7)
                sheet.delete_rows(1)
                sheet.insert_rows(2,2)
                sheet.delete_cols(11,2)
                total_par=sheet[1]
                profesor=sheet[6]
                profesor[0].value="Profesor"
                total_par[0].value=nombre_clase
                total_par[1].value="Total Participantes"
                total_par[2].value=int(participantes)-1
                total_aus=sheet[2]
                total_aus[1].value="Ausentes"
                total_aus[2].value=int(ausentes)
                total_est=sheet[3]
                total_est[1].value="Total Estudiantes"
                total_est[2].value=int(ausentes)+int(participantes)-1
                for n in range(1,int(invitados)):
                    sheet.cell(row = n+6, column = 1).value=n
                for row in sheet_Total.iter_rows(min_row=1,max_row=sheet.max_row,min_col=1, max_col=sheet.max_column):
                    for cell in row:
                               cell.value=0
                i=1
                for row in sheet.iter_rows():
                    j=0
                    fila=sheet_Total[i]
                    i=i+1
                    for cell in row:
                               fila[j].value=cell.value
                               if j==6:
                                   fila[j].number_format='mm-dd-yyyy'
                               j=j+1
                wb.save(filesheet)
            i=i+1
            driver.back()
            wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="webexIframeContainer"]')))
            driver.switch_to_frame(driver.find_element_by_xpath('//*[@id="webexIframeContainer"]'))
            wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td/form/table/tbody/tr[5]/td/table')))
            correos.append(correo)
            os.remove(os.path.join(filepath,correo+' '+nombre_clase+'.xlsx'))
            os.remove(os.path.join(filepath,newname))
        wb=openpyxl.load_workbook(filesheet)
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        wb.save(filesheet)
        self.correos=list(set(correos))
        self.ajustarTamaño(filesheet)
        self.ajustarTamaño(informeGeneral)
    
    def ajustarTamaño(self, filesheet):
        wb=openpyxl.load_workbook(filesheet)
        for n in wb.get_sheet_names():
             worksheet=wb.get_sheet_by_name(n)
             for column_cells in worksheet.columns:
                   length = max(len(self.as_text(cell.value)) for cell in column_cells)+2
                   worksheet.column_dimensions[column_cells[0].column_letter].width = length
                   for cell in column_cells:
                       cell.alignment = Alignment(horizontal='center')

        wb.save(filesheet)

    def as_text(self, value):
       if value is None:
           return ""
       return str(value)
       
    def tearDown(self):
        self.driver.close()

    def enviarCorreo(self, correo):
        msg = MIMEMultipart() 
        message = "Reporte Cisco Webex Training"
        password = "G849JNN43"
        msg['From'] = "csjulian@unicauca.edu.co"
        msg['To'] = correo
        msg['Subject'] = "Reporte de la fecha "+self.fecha+" Cisco Webex Training"
        carpeta=self.filesheet
        file=open(carpeta, 'rb')
        part=MIMEApplication(file.read())
        part['Content-Disposition']='attachment; filename='+basename(carpeta)
        msg.attach(part)
        msg.attach(MIMEText(message, 'plain'))
        server = smtplib.SMTP('smtp.gmail.com: 587')
        server.starttls()
        server.login(msg['From'], password)
        server.sendmail(msg['From'], msg['To'], msg.as_string())
        server.quit()
        print ("successfully sent email to %s:" % (msg['To']))

if __name__ == '__main__':
	  unittest.main()